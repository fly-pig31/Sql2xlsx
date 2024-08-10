import os
import openpyxl
from docx import Document

file = "111"
data_path = "resources/data3.xlsx"
result_path = f"result/result_{file}.xlsx"

# 打开原始 Excel 文件
workbook = openpyxl.load_workbook(data_path)

# 创建一个新的 Excel 文件
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# 获取所有工作表名
sheet_names = workbook.sheetnames

# 选择要读取的工作表
sheet = workbook[sheet_names[0]]  # 这里选择第一个工作表

# 标记当前行数
row_num = 0

# 获取 "resources\scanner" 文件夹下以 .docx 结尾的文件名
docx_files = [os.path.splitext(file)[0] for file in os.listdir(f"resources/{file}") if file.endswith(".docx")]


def read_docx_tables(doc):
    tables = []

    for table in doc.tables:
        table_data = []  # 创建一个内部列表来存储表格数据
        for row in table.rows:
            row_data = [cell.text for cell in row.cells]
            table_data.append(row_data)  # 将每行数据添加到内部列表中
        tables.append(table_data)  # 将内部列表添加到主列表中

    return tables


def read_vul_tables(vname: str, port_table: list):
    tables = []
    if len(port_table[0]) > 0:
        port_table[0].pop(0)  # 去掉第一行标题

    for rows in port_table[0]:
        if len(rows) > 3 and isinstance(rows[3], str):
            if vname in rows[3]:
                tables.append(rows)

    # 如果tables里的列表多于1个，则将列表合并
    if len(tables) > 1:
        merged_table = []
        for items in zip(*tables):
            merged_table.append('/'.join(items))

        tables = [merged_table]
    # tables[0].pop(3)
    if tables:
        return tables[0]
    return None


def get_table_data(key: str, my_table: list):
    for my_row in my_table:
        if my_row[0] == key:
            return my_row[1]
    return ""


docx_count = 0

# 遍历工作表的行
for row in sheet.iter_rows(values_only=True):
    new_row = list(row)
    if len(new_row) >= 7:  # 确保新行至少有7列数据

        # 删除第6列和第7列
        del new_row[5]  # 删除 F 列
        del new_row[5]  # 删除 G 列

        # 检查第三列数据是否是某个 .docx 文件名
        if new_row[3] in docx_files:
            docx_file_path = os.path.join(f"resources/{file}", new_row[3] + ".docx")
            if os.path.exists(docx_file_path):
                docx_count += 1
                doc = Document(docx_file_path)
                doc_tables = read_docx_tables(doc)
                vulnerabilities = []
                current_vulnerability = []  # 用于存储当前漏洞信息
                vul_tables = []
                port_tables = []
                for table in doc_tables:
                    for row in table:
                        if row[0] == "威胁分值":
                            if float(row[1]) > 4.0:
                                vul_tables.append(table)
                            else:
                                print(f"{new_row[3]} 的威胁分值小于 4.0，不处理")
                    if table[0] == ["端口", "协议", "服务", "漏洞"]:
                        port_tables.append(table)
                for table in vul_tables:
                    # print(table)
                    if current_vulnerability:  # 如果当前漏洞信息不为空，则添加到漏洞列表中
                        vulnerabilities.append(current_vulnerability)
                        current_vulnerability = []  # 重置当前漏洞信息列表
                    vulnerability_name = get_table_data("漏洞名称", table)

                    port = read_vul_tables(vulnerability_name, port_tables)

                    if port:
                        current_vulnerability.append(port[0])  # 添加解决办法
                        current_vulnerability.append(port[1])  # 添加解决办法
                        current_vulnerability.append(port[2])  # 添加解决办法
                    else:
                        print(f"未找到 {vulnerability_name} 的解决办法")
                        current_vulnerability.append("")  # 添加解决办法
                        current_vulnerability.append("")  # 添加解决办法
                        current_vulnerability.append("")  # 添加解决办法
                    current_vulnerability.append(vulnerability_name)  # 添加漏洞名称
                    current_vulnerability.append(get_table_data("解决办法", table))  # 添加解决办法

                    # 处理最后一个漏洞信息
                if current_vulnerability:
                    vulnerabilities.append(current_vulnerability)

                    # 在新行后面加入漏洞信息
                for vulnerability_info in vulnerabilities:
                    new_sheet.append(new_row + vulnerability_info)
            else:
                print(f"无法找到{docx_file_path}")
        else:

            print(f"{new_row[3]}.docx 文件不存在，不处理")
            pass
        new_sheet.append(new_row)

row_num += 1

# 保存新的 Excel 文件
new_workbook.save(result_path)

# 关闭 Excel 文件
workbook.close()
new_workbook.close()
